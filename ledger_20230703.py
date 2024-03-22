import email_parsing
import category_template

import openpyxl
from openpyxl.utils import get_column_letter
import datetime
from dateutil.parser import parse # for 날짜시간 format



def makeCSV(items):
    wb = openpyxl.Workbook()
    ws = wb.active

    lines = []  # category exception 추가
    lines_exception = []
    # date
    for item in items:
        if(type(item) == str):  # category exception 추가
            lines_exception.append([item, datetime.datetime(1970, 1, 1, 0, 0, 0), 0, '', '', '', ''])
            continue
        dateStr = item[1]
        date = item[1]
        if len(dateStr) == 11:  # 06.29 18:00
            year = datetime.datetime.now().year
            month = int(dateStr[0:2])
            day = int(dateStr[3:5])
            hour = int(dateStr[6:8])
            min = int(dateStr[9:11])
            sec = 0
            date = datetime.datetime(year, month, day, hour, min, sec)
        elif len(dateStr) == 14: # 23.06.29 18:00
            year = int(''.join({"20", dateStr[0:2]})) # 2023
            month = int(dateStr[3:5])
            day = int(dateStr[6:8])
            hour = int(dateStr[9:11])
            min = int(dateStr[12:14])
            sec = 0
            date = datetime.datetime(year, month, day, hour, min, sec)
        elif len(dateStr) == 16: # 2023.06.29 18:00
            year = int(dateStr[0:4])
            month = int(dateStr[5:7])
            day = int(dateStr[8:10])
            hour = int(dateStr[11:13])
            min = int(dateStr[14:16])
            sec = 0
            date = datetime.datetime(year, month, day, hour, min, sec)
        else:
            try:
                date = parse(item[1])
            except:
                date = datetime.datetime(1970, 1, 1, 0, 0, 0)
        item[1] = date

        amount = int(item[2].replace(',',''))
        item[2] = amount
        lines.append(item)

    # sort
    lines.sort(key = lambda x:x[1])
    lines.insert(0, ['입출', '날짜', '가격', '결제도구', '업체', '카테고리', '세부사항'])
    for item in lines_exception :  # category exception 추가
        lines.append(item)

    # border
    thin_border = openpyxl.styles.Border(left=openpyxl.styles.borders.Side(style='thin'), 
                     right=openpyxl.styles.borders.Side(style='thin'), 
                     top=openpyxl.styles.borders.Side(style='thin'), 
                     bottom=openpyxl.styles.borders.Side(style='thin'))

    # subtotal 정리용 year,month 숫자 추가
    lines[0].insert(2, 'year,month')
    for line in lines[1:]:
        date = line[1]
        year = date.year
        month = date.month
        yearmonth = f'{year}{month:02d}'
        line.insert(2, yearmonth)

    # inserCSV
    for row, item in enumerate(lines):
        row += 1
        for col, field in enumerate(item):
            col += 1
            ws.cell(row, col).value = field
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            if type(field) == datetime.datetime:
                ws.cell(row, col).number_format = 'yyyy-mm-dd hh:mm:ss' # date칸 표시형식 yyyy-mm-dd hh:mm:ss
            if type(field) == int:
                ws.cell(row, col).number_format = '#,###'

    # 열너비, set column width FAIL
    #for col in ws.columns:
    #    max_length = 0
    #    column = col[0].column
    #
    #    for cell in col:
    #        if len(str(cell.value)) > max_length:
    #            #max_length = len(str(cell.value))
    #            max_length = len(cell)
    #            print(', '.join([cell.coordinate, str(max_length)]))
    #
    #    adjusted_width = (max_length + 2) * 1.2
    #    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save("ledger.xlsx")

if __name__ == '__main__':
    try:
        infos = []
        count_item = 0
        email = email_parsing.Email()
        messages = email.messages
        #messages = messages[-1:]
        for msg in messages:
            #print(msg)
            #print('')
            items = msg.split('[Web발신]')
            items = items[1:]
            #print(len(items))

            if len(items) == 0 :    # [Web발신] 아닌 경우 중 단건
                lines = msg.split('\n')
                lines = lines[2:]
                items.append('\n'.join(lines))

            for item in items:
                count_item += 1
                item = item.strip()
                item = item.replace('\r\n', '\n')
                
                info = category_template.transform(item)
                
                if info == '':
                    print(item)
                    infos.append(item)
                #if len(info) > 3 and info[3] == '[현대카드] 자동납부':
                #    print(item)
                #    print(info)
                if info != '':
                    print(info)
                    infos.append(info)
        
        print(''.join(['count_item: ', str(count_item)]))        
        category_template.print_template()
        
        makeCSV(infos)
        
    except Exception as e:
        print('Exception', e)