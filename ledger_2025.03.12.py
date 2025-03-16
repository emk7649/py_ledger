import email_parsing
import category_template

import re
import openpyxl
import datetime
from dateutil.relativedelta import relativedelta
from dateutil.parser import parse # for 날짜시간 format

# vscode에서 현재경로 문제로 인해 추가
import os
folder_path = os.path.dirname(os.path.abspath(__file__))
os.chdir(folder_path)

def makeCSV(items):
    wb = openpyxl.Workbook()
    ws = wb.active

    lines = []  # category exception 추가
    lines_exception = []

    # exception 맨 밑에 넣기
    for item in items:
        if(type(item) == str):  # regex fail되어 msg str input인 경우, category exception 추가
            lines_exception.append([item, datetime.datetime(1970, 1, 1, 0, 0, 0), 0, '', '', '', ''])
        else:
            lines.append(item)

    # sort
    lines.sort(key = lambda x:x[1]) # 날짜순 sort
    lines.insert(0, ['입출', '날짜', '가격KRW', '가격currency', '결제도구', '업체', '카테고리', '세부사항'])
    for item in lines_exception :  # sort 후 category exception 추가
        lines.append(item)

    # border
    thin_border = openpyxl.styles.Border(left=openpyxl.styles.borders.Side(style='thin'), 
                     right=openpyxl.styles.borders.Side(style='thin'), 
                     top=openpyxl.styles.borders.Side(style='thin'), 
                     bottom=openpyxl.styles.borders.Side(style='thin'))

    # subtotal 정리용 year,month 숫자 추가
    lines[0].insert(2, 'year,month')
    for line in lines[1:]:
        date = line[1]
        year = date.year
        month = date.month
        yearmonth = f'{year}{month:02d}'
        line.insert(2, yearmonth)

    # inserCSV
    for row, item in enumerate(lines):
        row += 1
        for col, field in enumerate(item):
            col += 1
            ws.cell(row, col).value = field
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            if type(field) == datetime.datetime:
                ws.cell(row, col).number_format = 'yyyy-mm-dd hh:mm:ss' # date칸 표시형식 yyyy-mm-dd hh:mm:ss
            if type(field) == int:
                ws.cell(row, col).number_format = '#,###'

    # 열너비, set column width FAIL
    #for col in ws.columns:
    #    max_length = 0
    #    column = col[0].column
    #
    #    for cell in col:
    #        if len(str(cell.value)) > max_length:
    #            #max_length = len(str(cell.value))
    #            max_length = len(cell)
    #            print(', '.join([cell.coordinate, str(max_length)]))
    #
    #    adjusted_width = (max_length + 2) * 1.2
    #    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save("ledger.xlsx")

def processMessages2CsvData(messages):
    
    categoryTemplate = category_template.category_template()
    infos = []
    count_item = 0

    for msg in messages:
        year = -1
        
        msg = msg.replace("\r", "")
        msg = re.sub(r'\n+', '\n', msg) # 연속된 \n을 하나의 \n으로 치환

        ##
        regex = r'''\[Web발신\](\d\d\d\d)'''
        searchObj = re.search(regex, msg)
        if searchObj is not None:
            year = searchObj.groups()[0]
            lines = msg.split('\n')
            lines = lines[1:]
            item = '\n'.join(lines)
            items = []
            items.append(item)
        elif '[Web발신]' in msg :
            items = msg.split('[Web발신]')
            items = items[1:]
        else:
            items = []
            items.append(msg)

        for item in items:
            count_item += 1
            item = item.strip()
            item = item.replace('\r\n', '\n')
            
            record = categoryTemplate.processData2Record(item, int(year))
            
            if record == '':
                print(item)
                infos.append(item)
            else:
                print(record)
                infos.append(record)
    
    print(''.join(['count_item: ', str(count_item)]))
    categoryTemplate.print_template()
    return infos
    
    

if __name__ == '__main__':
    try:
        # make messages
        login = email_parsing.MailLogin()
        email = email_parsing.Email(login.mail_address, login.mail_password, login. mail_box)
        messages = email.messages

        infos = processMessages2CsvData(messages)
        makeCSV(infos)
        
    except Exception as e:
        print('Exception', e)
