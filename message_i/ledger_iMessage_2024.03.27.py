#import email_parsing
import category_template

import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import csv
import re

# vscode에서 현재경로 문제로 인해 추가
import os
folder_path = os.path.dirname(os.path.abspath(__file__))
os.chdir(folder_path)

def makeCSV(items):
    wb = openpyxl.Workbook()
    ws = wb.active

    lines = []  # category exception 추가
    lines_exception = []

    # exception 맨 밑에 넣기
    for item in items:
        if(type(item) == str):  # category exception 추가, re-matching fail되어 msg str input인 경우
            lines_exception.append([item, datetime.datetime(1970, 1, 1, 0, 0, 0), 0, '', '', '', ''])
            continue
        lines.append(item)

    # sort
    lines.sort(key = lambda x:x[1])
    lines.insert(0, ['입출', '날짜', '가격KRW', '가격currency', '결제도구', '업체', '카테고리', '세부사항'])
    for item in lines_exception :  # category exception 추가
        lines.append(item)

    # border
    thin_border = openpyxl.styles.Border(left=openpyxl.styles.borders.Side(style='thin'), 
                     right=openpyxl.styles.borders.Side(style='thin'), 
                     top=openpyxl.styles.borders.Side(style='thin'), 
                     bottom=openpyxl.styles.borders.Side(style='thin'))

    # subtotal 정리용 year,month 숫자 추가
    lines[0].insert(2, 'year,month')
    for line in lines[1:]:
        date = line[1]
        year = date.year
        month = date.month
        yearmonth = f'{year}{month:02d}'
        line.insert(2, yearmonth)

    # inserCSV
    for row, item in enumerate(lines):
        row += 1
        for col, field in enumerate(item):
            col += 1
            ws.cell(row, col).value = field
            ws.cell(row, col).border = thin_border
            ws.cell(row, col).alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            if type(field) == datetime.datetime:
                ws.cell(row, col).number_format = 'yyyy-mm-dd hh:mm:ss' # date칸 표시형식 yyyy-mm-dd hh:mm:ss
            if type(field) == int:
                ws.cell(row, col).number_format = '#,###'

    # 열너비, set column width FAIL
    #for col in ws.columns:
    #    max_length = 0
    #    column = col[0].column
    #
    #    for cell in col:
    #        if len(str(cell.value)) > max_length:
    #            #max_length = len(str(cell.value))
    #            max_length = len(cell)
    #            print(', '.join([cell.coordinate, str(max_length)]))
    #
    #    adjusted_width = (max_length + 2) * 1.2
    #    ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save("ledger.xlsx")

if __name__ == '__main__':
    messages = []
    years = []

    csv_files = ['신한.csv', '우리.csv', '현대.csv']
    for file in csv_files:
        if not os.path.exists(file):
            continue
        f = open(file,'r', encoding='UTF8')
        rdr = csv.reader(f)
        for i, line in enumerate(rdr):
            if(len(line) < 5):
                continue

            data = line[0]
            regex = r'''\d\d\d\d'''
            searchObj = re.search(regex, data)
            if searchObj is not None:
                # inline 중에도 사용할 수 있도록 [Web발신] 옆에 str_year 붙이기
                #years.append(searchObj.group())
                year = searchObj.group()
                line[4] = line[4].replace('[Web발신]', f'[Web발신]{year}')
                messages.append(line[4])
        f.close()

    try:
        categoryTemplate = category_template.category_template()
        infos = []
        count_item = 0
        #messages = messages[-1:]
        for msg in messages:
            #print(msg)
            #print('')

            # 하나의 메시지는 하나의 Item
            # [Web발신] 있으면 \n까지 지우기
            # [Web발신]\d\d\d\d 있으면 (\d\d\d\d) -> year 치환
            lines = msg.split('\n')
            if len(lines) == 0 :
                continue
            if '[Web발신]' in lines[0] :
                regex = r'''\[Web발신\](\d\d\d\d)'''
                searchObj = re.search(regex, lines[0])
                if searchObj is not None:
                    year = searchObj.groups()[0]
                    lines = lines[1:]
                else:
                    lines = lines[0:]
                item = '\n'.join(lines)

                count_item += 1
                item = item.strip()
                item = item.replace('\r\n', '\n')
                
                record = categoryTemplate.processData2Record(item, int(year))
                
                #re-matching fail:msg str, re-matching success:list_sorted
                if record == '':
                    print(item)
                    infos.append(item)
                if record != '':
                    print(record)
                    infos.append(record)
        
        print(''.join(['count_item: ', str(count_item)]))        
        categoryTemplate.print_template()
        
        makeCSV(infos)
        
    except Exception as e:
        print('Exception', e)